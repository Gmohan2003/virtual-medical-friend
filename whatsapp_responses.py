# Python program to read an excel file

# import openpyxl module
import openpyxl
def responses(input_message):
    # Give the location of the file
    path = "hack 2022.xlsx"
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active

    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or
    # column integer is 1, not 0.

    # Cell object is created by using
    # sheet object's cell() method.

    for i in range(1, sheet_obj.max_row+1):
        cell_obj = sheet_obj.cell(row=i, column = 1)
        if(cell_obj.value == input_message.lower()):
            cell_obj2= sheet_obj.cell(row=i, column = 2)
            return cell_obj2.value
    return "Invalid Input...Please Enter Help!"
    # Print value of cell object
    # using the value attribute

